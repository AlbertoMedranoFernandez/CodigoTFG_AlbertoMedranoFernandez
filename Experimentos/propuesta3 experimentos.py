import cv2
import mediapipe as mp
import math
import time
import os
import openpyxl

def main():
    row = 3
    media_ta_parpadeos = 0
    media_ta_micro = 0
    ubi = 'C:/Users/MSI/Desktop/bbdd/'

    libro = openpyxl.load_workbook('Experimentos.xlsx')
    hoja = libro['Propuesta3']

    videos = os.listdir(ubi)
    for video in videos:
        contParpadeos, contMicros = detector(video, ubi)
        hoja.cell(row, 1, video)  # Escribir nombre video
        hoja.cell(row, 4, contParpadeos)  # Escribir parpadeos
        hoja.cell(row, 5, contMicros)  # Escribir micro sueños

        manual_parpadeos = hoja.cell(row, 2).value  # Obtener número manual parpadeos
        algoritmo_parpadeos = contParpadeos
        manual_micro = hoja.cell(row, 3).value  # Obtener número manual micro sueños
        algoritmo_micro = contMicros

        dif_parpadeos = abs(manual_parpadeos - algoritmo_parpadeos)  # Error parpadeos
        dif_micro = abs(manual_micro - algoritmo_micro)  # Error micro sueños

        # Tasa de acierto de parpadeos
        if manual_parpadeos > algoritmo_parpadeos:
            ta_parpadeos = 1 - (dif_parpadeos / manual_parpadeos)
        else:
            ta_parpadeos = 1 - (dif_parpadeos / algoritmo_parpadeos)
        media_ta_parpadeos += ta_parpadeos

        # Tasa de acierto de micro sueños
        if manual_micro > algoritmo_micro:
            ta_micro = 1 - (dif_micro / manual_micro)
        else:
            ta_micro = 1 - (dif_micro / algoritmo_micro)
        media_ta_micro += ta_micro

        por_parpadeos = ta_parpadeos * 100  # Porcentaje de parpadeos
        por_micro = ta_micro * 100  # Porcentaje de micro sueños

        # Escribir resultados
        hoja.cell(row, 6, ta_parpadeos)
        hoja.cell(row, 7, por_parpadeos)
        hoja.cell(row, 8, ta_micro)
        hoja.cell(row, 9, por_micro)

        row += 1

    media_ta_parpadeos = media_ta_parpadeos / len(videos)  # Media tasas de aciertos de parpadeos
    media_por_parpadeos = media_ta_parpadeos * 100  # Media porcentajes de parpadeos
    media_ta_micro = media_ta_micro / len(videos)  # Media tasas de aciertos de micro sueños
    media_por_micro = media_ta_micro * 100  # Media porcentajes de micro sueños

    # Escribir resultados
    hoja.cell(row, 6, media_ta_parpadeos)
    hoja.cell(row, 7, media_por_parpadeos)
    hoja.cell(row, 8, media_ta_micro)
    hoja.cell(row, 9, media_por_micro)

    libro.save('Experimentos.xlsx')


def detector(nombre, ubi):
    # Variables
    parpadeo = False
    contParpadeos = 0
    tiempo = 0
    inicio = 0
    final = 0
    contSueños = 0

    calibrado = False
    puntos = []

    # Encender cámara
    video = ubi + nombre
    cap = cv2.VideoCapture(video)
    cap.set(3, 1280)
    cap.set(4, 720)

    # Para dibujar malla facial
    drawingSpec = mp.solutions.drawing_utils.DrawingSpec(thickness=1, circle_radius=1)

    # Para crear la malla facial
    faceMesh = mp.solutions.face_mesh.FaceMesh(max_num_faces=1, refine_landmarks=True)

    # Bucle principal
    while True:

        ret, frame = cap.read()
        if not ret:  # Acabar si no quedan fotogramas
            break
        frameRGB = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)  # Corrección de color

        resultados = faceMesh.process(frameRGB)

        if resultados.multi_face_landmarks:  # Si se detecta alguna cara
            for rostros in resultados.multi_face_landmarks:  # Mostrar cara detectada
                mp.solutions.drawing_utils.draw_landmarks(frame, rostros, mp.solutions.face_mesh.FACEMESH_IRISES, None,
                                                          drawingSpec)

                # Extraer los puntos de la cara detectada
                lista = extraerPuntos(frame, rostros.landmark)

                # Ojo derecho
                longitud1 = calcularDistancia(lista[470], lista[472])

                # Ojo izquierdo
                longitud2 = calcularDistancia(lista[475], lista[477])

                # Frente y barbilla
                longitud3 = calcularDistancia(lista[10], lista[152])

                # Calibrado automatico
                if calibrado == False:
                    cv2.putText(frame, f'Calibrando...', (75, 60), cv2.FONT_HERSHEY_SIMPLEX, 2, (255, 0, 0), 3)
                    if len(puntos) < 2:
                        aux = (longitud1 + longitud2) / 2
                        puntos.append([longitud3, aux])
                    else:
                        print(puntos)
                        pendiente = calcularPendiente(puntos[0], puntos[1])
                        if pendiente < 0.09 and pendiente > 0.045:
                            calibrado = True
                            print("Calibración existosa")
                        else:
                            puntos = []
                            print("Fallo calibración")

                # Cuando ya está calibrado
                else:
                    # Calcular umbral
                    x, y = puntos[0][0], puntos[0][1]
                    umbral = (pendiente * (longitud3 - x) + y)

                    # Controlar inicio y final del parpadeo
                    print(longitud3, longitud1, longitud2, umbral)
                    parpadeo, finalParpadeo, inicio, final = parpadear(longitud1, longitud2, parpadeo, inicio, final, umbral)

                    # Comporbar si ha sido micro sueño
                    if finalParpadeo == True:
                        tiempo, contParpadeos, contSueños, finalParpadeo = esMicroSueño(inicio, final, contParpadeos, contSueños)

                    # Mostrar datos
                    mostrarDatos(frame, contParpadeos, contSueños, tiempo)

        cv2.imshow('Frame', frame)
        key = cv2.waitKey(25) & 0xFF

        # Pulsar "q" para acabar
        if key == ord("q"):
            break

    # Cerrar la cámara
    cap.release()
    cv2.destroyAllWindows()

    return contParpadeos, contSueños


# Función para calcular la pendiente de la ecuación lineal con la que se calcula el umbral
def calcularPendiente(punto1, punto2):
    try:
        print('pendiente: ', (punto2[1] - punto1[1]) / (punto2[0] - punto1[0]))
        return abs((punto2[1] - punto1[1]) / (punto2[0] - punto1[0]))
    except ZeroDivisionError:
        return 0

# Función para pasar las coordenadas de los puntos de referencia a píxeles
def extraerPuntos(frame, landmarks):
    lista = []

    for id in range(len(landmarks)):
        al, an, c = frame.shape  # alto, ancho, canales
        x, y = int(landmarks[id].x * an), int(landmarks[id].y * al)
        lista.append([x, y])
        if len(lista) == 478:
            return lista

# Función para calcular la distancia entre dos puntos de referencia
def calcularDistancia(punto1, punto2):
    return math.hypot(punto2[0] - punto1[0], punto2[1] - punto1[1])

# Función para controlar cuando se abren y cierran los ojos
def parpadear(longitud1, longitud2, parpadeo, inicio, final, umbral):

    finalParpadeo = False

    if longitud1 <= umbral and longitud2 <= umbral and parpadeo == False:
        parpadeo = True
        inicio = time.time()

    if longitud1 > umbral and longitud2 > umbral and parpadeo == True:
        parpadeo = False
        final = time.time()
        finalParpadeo = True

    return parpadeo, finalParpadeo, inicio, final

# Función para determinar si se ha producido un parpadeo o un micro sueño
def esMicroSueño(inicio, final, contParpadeos, contSueños):
    # Tiempo entre que se abre y cierra el ojo
    tiempo = round(final - inicio, 1)

    # Contador de parpadeos
    if tiempo < 1:
        contParpadeos = contParpadeos + 1

    # Contador de micro Sueños
    elif tiempo >= 1:
        contSueños = contSueños + 1

    finalParpadeo = False

    return tiempo, contParpadeos, contSueños, finalParpadeo

# Función para mostrar en la ventana el número de parpadeos y de micro sueños y el tiempo
def mostrarDatos(frame, contParpadeos, contSueños, tiempo):
    cv2.putText(frame, f'Parpadeos: {int(contParpadeos)}', (75, 60), cv2.FONT_HERSHEY_SIMPLEX, 1,
                (0, 255, 0), 3)
    cv2.putText(frame, f'Micro sueños: {int(contSueños)}', (75, 100), cv2.FONT_HERSHEY_SIMPLEX, 1,
                (0, 0, 255), 3)
    cv2.putText(frame, f'Duracion: {str(tiempo)}', (75, 140), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0),
                3)


if __name__ == "__main__":
    main()